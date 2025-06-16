from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from .models import Przychod, Rozchod, Plyta
from datetime import datetime
from django.utils.dateparse import parse_date
from moneyed import Money, EUR

def ExportAll(request, rok, sw):
    if sw == '1':
        start_d = "2023-01-01"
        stop_d = "2023-12-31"
        fl = 0
    elif sw == '2':
        start_d = "2024-01-01"
        stop_d = "2024-12-31"
        fl = 0
    else:
        start_d = "2022-01-01"
        stop_d = "2024-12-31"
        fl = 1

    return ExportByDateRange(request, start_d, stop_d, fl)


def ExportByDateRange(request, start_d, stop_d, fl):
    # Pobranie parametrów GET
    start_date_str = start_d
    end_date_str = stop_d

    # Weryfikacja obecności obu dat
    if not start_date_str or not end_date_str:
        return HttpResponse("Błąd: Brak jednej z dat. Podaj oba parametry: start_date i end_date.", status=400)

    # Parsowanie dat
    try:
        start_date = parse_date(start_date_str)
        end_date = parse_date(end_date_str)
        if not start_date or not end_date:
            raise ValueError("Niepoprawny format daty.")
    except ValueError:
        return HttpResponse("Błąd: Niepoprawny format dat. Użyj formatu RRRR-MM-DD.", status=400)

    # Filtrowanie danych w przedziale dat
    przychody = Przychod.objects.filter(data__range=(start_date, end_date))
    rozchody = Rozchod.objects.filter(data__range=(start_date, end_date))
    plyty = Plyta.objects.filter(rodzaj='dre').order_by('magazyn')

    # Jeśli brak danych, zwracamy komunikat
    if not przychody.exists() and not rozchody.exists():
        return HttpResponse("Brak danych w podanym przedziale dat.", status=404)

    # Tworzenie workbooka
    wb = Workbook()
    ws = wb.active
    if fl == 0:
        ws.title = f"Płyty i operacje od {start_date} do {end_date}"
    else:
        ws.title = f"Płyty i operacje - całość"

    # Styl nagłówka
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="a04000", end_color="a04000", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Styl danych
    data_alignment = Alignment(horizontal="center", vertical="center")
    currency_alignment = Alignment(horizontal="right", vertical="center")
    plyta_fill = PatternFill(start_color="fcf3cf", end_color="fcf3cf", fill_type="solid")  # Blado żółty

    # Styl obramowania
    border_color = "e59866"
    thin_border = Border(
        left=Side(style="thin", color=border_color),
        right=Side(style="thin", color=border_color),
        top=Side(style="thin", color=border_color),
        bottom=Side(style="thin", color=border_color)
    )

    # Nagłówki
    headers = [
        "ID Prod.", "Magazyn", "Nazwa", "Stan", "Jednostka", "Opis", "Cena",
        "Typ Operacji", "ID Dok.", "Źródło/Cel", "Data", "Ilość", "Jednostka", "Kwota/Cena", "SDE"
    ]
    ws.append(headers)

    # Formatowanie nagłówków
    for col_num, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Dane
    for plyta in plyty:
        mag = ''
        if plyta.magazyn == 'MAGAZYN1':
            mag = 'SZPARAGOWA'
        if plyta.magazyn == 'MAGAZYN2':
            mag = 'PODOLANY'
        if plyta.magazyn == 'MAGAZYN3':
            mag = 'MAG. U DOST.'

        # Dane Plyta
        plyta_data = [
            plyta.prod_id,
            mag,
            plyta.nazwa,
            plyta.stan,
            plyta.jm,
            plyta.opis,
            f"{plyta.cena.amount:.2f}".replace('.', ',') if plyta.cena else ""
            #f"{plyta.cena.amount} {plyta.cena.currency}" if plyta.cena else ""
        ]

        # Dodajemy linie z tabeli Przychod powiązane z tą Plyta
        plyta_przychody = przychody.filter(plyta=plyta)
        for przychod in plyta_przychody:
            przychod_row = plyta_data + [
                "Przychód",
                przychod.doc_id,
                przychod.zrodlo,
                przychod.data,
                przychod.ilosc,
                przychod.jm,
                f"{przychod.cena.amount:.2f}".replace('.', ',') if przychod.cena else "",
                # f"{przychod.cena.amount} {przychod.cena.currency}" if przychod.cena else "",
                " - "
            ]
            ws.append(przychod_row)

            # Formatowanie wiersza Przychod
            row_num = ws.max_row
            for col_num, value in enumerate(przychod_row, start=1):
                cell = ws.cell(row=row_num, column=col_num)
                if col_num == 7 or col_num == 14:
                    cell.alignment = currency_alignment
                else:
                    cell.alignment = data_alignment
                if col_num in [1, 2, 3, 4, 5, 6, 7]:
                    cell.fill = plyta_fill
                cell.border = thin_border

        # Dodajemy linie z tabeli Rozchod powiązane z tą Plyta
        plyta_rozchody = rozchody.filter(plyta=plyta)
        for rozchod in plyta_rozchody:
            rozchod_row = plyta_data + [
                "Rozchód",
                rozchod.doc_id,
                rozchod.cel,
                rozchod.data,
                rozchod.ilosc,
                rozchod.jm,
                f"{rozchod.kwota.amount:.2f}".replace('.', ',') if rozchod.kwota else "",
                #f"{rozchod.kwota.amount} {rozchod.kwota.currency}" if rozchod.kwota else "",
                f"{rozchod.nr_sde.nazwa}" if rozchod.nr_sde else ""
            ]
            ws.append(rozchod_row)

            # Formatowanie wiersza Rozchod
            row_num = ws.max_row
            for col_num, value in enumerate(rozchod_row, start=1):
                cell = ws.cell(row=row_num, column=col_num)
                if col_num == 7 or col_num == 14:
                    cell.alignment = currency_alignment
                else:
                    cell.alignment = data_alignment
                if col_num in [1, 2, 3, 4, 5, 6, 7]:
                    cell.fill = plyta_fill
                cell.border = thin_border

    # Automatyczne dopasowanie szerokości kolumn
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Pobranie litery kolumny
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

    # Zapisanie workbooka do strumienia
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    file_name = ''
    if fl == 0:
        file_name = "Zestawienie_"+str(start_date)+"_do_"+str(end_date)+".xlsx"
    else:
        file_name = "Zestawienie_calosc.xlsx"

    # Przygotowanie odpowiedzi HTTP
    response = HttpResponse(output, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{file_name}"'
    return response


def ExportByDateRangeAndElement(request, start_d, stop_d, pk, mag, fl, sel):
    # Pobranie parametrów GET
    start_date_str = start_d
    end_date_str = stop_d

    # Weryfikacja obecności obu dat
    if not start_date_str or not end_date_str:
        return HttpResponse("Błąd: Brak jednej z dat. Podaj oba parametry: start_date i end_date.", status=400)

    # Parsowanie dat
    try:
        start_date = parse_date(start_date_str)
        end_date = parse_date(end_date_str)
        if not start_date or not end_date:
            raise ValueError("Niepoprawny format daty.")
    except ValueError:
        return HttpResponse("Błąd: Niepoprawny format dat. Użyj formatu RRRR-MM-DD.", status=400)

    # Filtrowanie danych w przedziale dat
    przychody = Przychod.objects.filter(data__range=(start_date, end_date))
    rozchody = Rozchod.objects.filter(data__range=(start_date, end_date))
    plyty = Plyta.objects.filter(pk=pk, rodzaj='dre').order_by('magazyn')

    # Jeśli brak danych, zwracamy komunikat
    if not przychody.exists() and not rozchody.exists():
        return HttpResponse("Brak danych w podanym przedziale dat.", status=404)

    # Tworzenie workbooka
    wb = Workbook()
    ws = wb.active
    if fl == 0:
        ws.title = f"Płyty i operacje od {start_date} do {end_date}"
    else:
        ws.title = f"Płyty i operacje - całość"

    # Styl nagłówka
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="a04000", end_color="a04000", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Styl danych
    data_alignment = Alignment(horizontal="center", vertical="center")
    currency_alignment = Alignment(horizontal="right", vertical="center")
    plyta_fill = PatternFill(start_color="fcf3cf", end_color="fcf3cf", fill_type="solid")  # Blado żółty

    # Styl obramowania
    border_color = "e59866"
    thin_border = Border(
        left=Side(style="thin", color=border_color),
        right=Side(style="thin", color=border_color),
        top=Side(style="thin", color=border_color),
        bottom=Side(style="thin", color=border_color)
    )

    # Nagłówki
    headers = [
        "ID Prod.", "Magazyn", "Nazwa", "Stan", "Jednostka", "Opis", "Cena",
        "Typ Operacji", "ID Dok.", "Źródło/Cel", "Data", "Ilość", "Jednostka", "Kwota/Cena", "SDE"
    ]
    ws.append(headers)

    # Formatowanie nagłówków
    for col_num, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Dane
    for plyta in plyty:
        mag = ''
        if plyta.magazyn == 'MAGAZYN1':
            mag = 'SZPARAGOWA'
        if plyta.magazyn == 'MAGAZYN2':
            mag = 'PODOLANY'
        if plyta.magazyn == 'MAGAZYN3':
            mag = 'MAG. U DOST.'

        # Dane Plyta
        plyta_data = [
            plyta.prod_id,
            mag,
            plyta.nazwa,
            plyta.stan,
            plyta.jm,
            plyta.opis,
            f"{plyta.cena.amount:.2f}".replace('.', ',') if plyta.cena else ""
        ]

        tfile = plyta.nazwa  #f"{plyta.nazwa}".replace('(', '_').replace(')', '_')

        # Dodajemy linie z tabeli Przychod powiązane z tą Plyta
        plyta_przychody = przychody.filter(plyta=plyta)
        for przychod in plyta_przychody:
            przychod_row = plyta_data + [
                "Przychód",
                przychod.doc_id,
                przychod.zrodlo,
                przychod.data,
                przychod.ilosc,
                przychod.jm,
                f"{przychod.cena.amount:.2f}".replace('.', ',') if przychod.cena else "",
                " - "
            ]
            ws.append(przychod_row)

            # Formatowanie wiersza Przychod
            row_num = ws.max_row
            for col_num, value in enumerate(przychod_row, start=1):
                cell = ws.cell(row=row_num, column=col_num)
                if col_num == 7 or col_num == 14:
                    cell.alignment = currency_alignment
                else:
                    cell.alignment = data_alignment
                if col_num in [1, 2, 3, 4, 5, 6, 7]:
                    cell.fill = plyta_fill
                cell.border = thin_border

        # Dodajemy linie z tabeli Rozchod powiązane z tą Plyta
        plyta_rozchody = rozchody.filter(plyta=plyta)
        for rozchod in plyta_rozchody:
            rozchod_row = plyta_data + [
                "Rozchód",
                rozchod.doc_id,
                rozchod.cel,
                rozchod.data,
                rozchod.ilosc,
                rozchod.jm,
                f"{rozchod.kwota.amount:.2f}".replace('.', ',') if rozchod.kwota else "",
                #f"{rozchod.kwota.amount} {rozchod.kwota.currency}" if rozchod.kwota else "",
                f"{rozchod.nr_sde.nazwa}" if rozchod.nr_sde else ""
            ]
            ws.append(rozchod_row)

            # Formatowanie wiersza Rozchod
            row_num = ws.max_row
            for col_num, value in enumerate(rozchod_row, start=1):
                cell = ws.cell(row=row_num, column=col_num)
                if col_num == 7 or col_num == 14:
                    cell.alignment = currency_alignment
                else:
                    cell.alignment = data_alignment
                if col_num in [1, 2, 3, 4, 5, 6, 7]:
                    cell.fill = plyta_fill
                cell.border = thin_border

    # Automatyczne dopasowanie szerokości kolumn
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Pobranie litery kolumny
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

    # Zapisanie workbooka do strumienia
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    file_name = "Raport_"+str(tfile)+".xlsx"

    # Przygotowanie odpowiedzi HTTP
    response = HttpResponse(output, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{file_name}"'
    return response


def inw_xls(request, mag, db):
    zero = Money('00.00', EUR)
    if mag == 'mag1':
        tytul = 'Inwentura magazynu Szparagowa'
    elif mag == 'mag2':
        tytul = 'Inwentura magazynu Podolany'
    elif mag == 'mag3':
        tytul = 'Inwentura u Dostawcy'
    else:
        print("MAGAZYN SPOZA ZAKRESU !!!")

    # Tworzenie workbooka
    wb = Workbook()
    ws = wb.active

    ws.title = tytul

    # Styl nagłówka
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="a04000", end_color="a04000", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Styl danych
    data_alignment = Alignment(horizontal="center", vertical="center")
    currency_alignment = Alignment(horizontal="right", vertical="center")
    plyta_fill = PatternFill(start_color="fcf3cf", end_color="fcf3cf", fill_type="solid")  # Blado żółty

    # Styl obramowania
    border_color = "e59866"
    thin_border = Border(
        left=Side(style="thin", color=border_color),
        right=Side(style="thin", color=border_color),
        top=Side(style="thin", color=border_color),
        bottom=Side(style="thin", color=border_color)
    )

    # Nagłówki
    headers = [
        "ID Prod.", "Magazyn", "Nazwa", "Opis", "Wartość", "Stan", "Brak ceny jedn."
    ]
    ws.append(headers)

    # Formatowanie nagłówków
    for col_num, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Dane
    for plyta in db:
        magw = ''
        if mag == 'mag1':
            magw = 'SZPARAGOWA'
        if mag == 'mag2':
            magw = 'PODOLANY'
        if mag == 'mag3':
            magw = 'MAG. U DOST.'

        bc = ''
        if plyta['brak_ceny'] != 0:
            bc = str(plyta['brak_ceny'])

        plyta_data = [
            plyta['id'],
            magw,
            plyta['nazwa'],
            plyta['opis'],
            f"{plyta['suma_wart'].amount:.2f}".replace('.', ',') if plyta['suma_wart'] else f"{zero.amount:.2f}".replace('.', ','),
            plyta['stan'],
            bc
        ]
        ws.append(plyta_data)

        # Formatowanie wierszy
        row_num = ws.max_row
        for col_num, value in enumerate(plyta_data, start=1):
            cell = ws.cell(row=row_num, column=col_num)
            if col_num == 5:
                cell.alignment = currency_alignment
            else:
                cell.alignment = data_alignment
            if col_num in [1, 2, 3, 4, 5, 6, 7]:
                cell.fill = plyta_fill
            cell.border = thin_border

    # Automatyczne dopasowanie szerokości kolumn
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Pobranie litery kolumny
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

    # Zapisanie workbooka do strumienia
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    file_name = tytul + ".xlsx"

    # Przygotowanie odpowiedzi HTTP
    response = HttpResponse(output, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{file_name}"'
    return response





















































































# def ExportAll(request, rok):
#     try:
#         year = int(rok)
#     except ValueError:
#         raise ValueError("Nieprawidłowy rok.")
#
#     # Filtrowanie tabeli Plyta: tylko rodzaj 'dre', sortowanie po magazyn
#     plyty = Plyta.objects.filter(rodzaj='dre').order_by('magazyn')
#
#     # Filtrowanie tabel Przychod i Rozchod dla danego roku
#     przychody = Przychod.objects.filter(data__year=year)
#     rozchody = Rozchod.objects.filter(data__year=year)
#
#     # Tworzenie workbooka
#     wb = Workbook()
#     ws = wb.active
#     ws.title = f"Płyty i operacje {year}"
#
#     # Styl nagłówka
#     header_font = Font(bold=True)
#     header_fill = PatternFill(start_color="a04000", end_color="a04000", fill_type="solid")
#     header_alignment = Alignment(horizontal="center", vertical="center")
#
#     # Styl danych
#     data_alignment = Alignment(horizontal="center", vertical="center")
#     currency_alignment = Alignment(horizontal="right", vertical="center")
#     plyta_fill = PatternFill(start_color="fcf3cf", end_color="fcf3cf", fill_type="solid")  # Blado żółty
#
#     # Styl obramowania
#     border_color = "e59866"
#     thin_border = Border(
#         left=Side(style="thin", color=border_color),
#         right=Side(style="thin", color=border_color),
#         top=Side(style="thin", color=border_color),
#         bottom=Side(style="thin", color=border_color)
#     )
#
#     # Nagłówki
#     headers = [
#         "ID Prod.", "Magazyn", "Nazwa", "Stan", "Jednostka", "Opis", "Cena",
#         "Typ Operacji", "ID Dok.", "Źródło/Cel", "Data", "Ilość", "Jednostka", "Kwota/Cena"
#     ]
#     ws.append(headers)
#
#     # Formatowanie nagłówków
#     for col_num, _ in enumerate(headers, start=1):
#         cell = ws.cell(row=1, column=col_num)
#         cell.font = header_font
#         cell.fill = header_fill
#         cell.alignment = header_alignment
#         cell.border = thin_border
#
#     # Zwiększenie wysokości wiersza nagłówka
#     ws.row_dimensions[1].height = 22.5  # 50% więcej (standard to 15)
#
#     # Dane
#     for plyta in plyty:
#         # Dodajemy główną linię dla Plyta
#         mag = ''
#         if plyta.magazyn == 'MAGAZYN1':
#             mag = 'SZPARAGOWA'
#         if plyta.magazyn == 'MAGAZYN2':
#             mag = 'PODOLANY'
#
#         plyta_row = [
#             plyta.prod_id,
#             mag,
#             plyta.nazwa,
#             plyta.stan,
#             plyta.jm,
#             plyta.opis,
#             f"{plyta.cena.amount} {plyta.cena.currency}" if plyta.cena else "",
#             "", "", "", "", "", "", ""
#         ]
#         ws.append(plyta_row)
#
#         # Formatowanie wiersza Plyta
#         row_num = ws.max_row
#         for col_num, value in enumerate(plyta_row, start=1):
#             cell = ws.cell(row=row_num, column=col_num)
#             # Kolumny 7 i 14: wyrównanie do prawej
#             if col_num == 7 or col_num == 14:
#                 cell.alignment = currency_alignment
#             else:
#                 cell.alignment = data_alignment
#             if col_num == 3:  # Kolumna "Nazwa" - tekst pogrubiony
#                 cell.font = Font(bold=True)
#             cell.fill = plyta_fill
#             cell.border = thin_border
#
#         # Dodajemy linie z tabeli Przychod powiązane z tą Plyta
#         plyta_przychody = przychody.filter(plyta=plyta)
#         for przychod in plyta_przychody:
#             przychod_row = [
#                 "", "", "", "", "", "", "",
#                 "Przychód",
#                 przychod.doc_id,
#                 przychod.zrodlo,
#                 przychod.data,
#                 przychod.ilosc,
#                 przychod.jm,
#                 f"{przychod.cena.amount} {przychod.cena.currency}" if przychod.cena else ""
#             ]
#             ws.append(przychod_row)
#
#             # Formatowanie wiersza Przychod
#             row_num = ws.max_row
#             for col_num, value in enumerate(przychod_row, start=1):
#                 cell = ws.cell(row=row_num, column=col_num)
#                 # Kolumny 7 i 14: wyrównanie do prawej
#                 if col_num == 7 or col_num == 14:
#                     cell.alignment = currency_alignment
#                 else:
#                     cell.alignment = data_alignment
#                 cell.border = thin_border
#
#         # Dodajemy linie z tabeli Rozchod powiązane z tą Plyta
#         plyta_rozchody = rozchody.filter(plyta=plyta)
#         for rozchod in plyta_rozchody:
#             rozchod_row = [
#                 "", "", "", "", "", "", "",
#                 "Rozchód",
#                 rozchod.doc_id,
#                 rozchod.cel,
#                 rozchod.data,
#                 rozchod.ilosc,
#                 rozchod.jm,
#                 f"{rozchod.kwota.amount} {rozchod.kwota.currency}" if rozchod.kwota else ""
#             ]
#             ws.append(rozchod_row)
#
#             # Formatowanie wiersza Rozchod
#             row_num = ws.max_row
#             for col_num, value in enumerate(rozchod_row, start=1):
#                 cell = ws.cell(row=row_num, column=col_num)
#                 # Kolumny 7 i 14: wyrównanie do prawej
#                 if col_num == 7 or col_num == 14:
#                     cell.alignment = currency_alignment
#                 else:
#                     cell.alignment = data_alignment
#                 cell.border = thin_border
#
#     # Automatyczne dopasowanie szerokości kolumn
#     for col in ws.columns:
#         max_length = 0
#         col_letter = col[0].column_letter  # Pobranie litery kolumny
#         for cell in col:
#             try:
#                 if cell.value:
#                     max_length = max(max_length, len(str(cell.value)))
#             except:
#                 pass
#         adjusted_width = max_length + 2  # Dodatkowy margines
#         ws.column_dimensions[col_letter].width = adjusted_width
#
#     # Zapisanie workbooka do strumienia
#     output = BytesIO()
#     wb.save(output)
#     output.seek(0)
#
#     # Przygotowanie odpowiedzi HTTP
#     response = HttpResponse(output, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
#     response["Content-Disposition"] = f'attachment; filename="Eksport_{year}.xlsx"'
#     return response

#
# def ExportByDateRange(request, start_d, stop_d):
#     # Pobranie parametrów GET
#     start_date_str = start_d
#     end_date_str = stop_d
#
#     # Weryfikacja obecności obu dat
#     if not start_date_str or not end_date_str:
#         return HttpResponse("Błąd: Brak jednej z dat. Podaj oba parametry: start_date i end_date.", status=400)
#
#     # Parsowanie dat
#     try:
#         start_date = parse_date(start_date_str)
#         end_date = parse_date(end_date_str)
#         if not start_date or not end_date:
#             raise ValueError("Niepoprawny format daty.")
#     except ValueError:
#         return HttpResponse("Błąd: Niepoprawny format dat. Użyj formatu RRRR-MM-DD.", status=400)
#
#     # Filtrowanie danych w przedziale dat
#     przychody = Przychod.objects.filter(data__range=(start_date, end_date))
#     rozchody = Rozchod.objects.filter(data__range=(start_date, end_date))
#     plyty = Plyta.objects.filter(rodzaj='dre').order_by('magazyn')
#
#     # Jeśli brak danych, zwracamy komunikat
#     if not przychody.exists() and not rozchody.exists():
#         return HttpResponse("Brak danych w podanym przedziale dat.", status=404)
#
#     # Tworzenie workbooka
#     wb = Workbook()
#     ws = wb.active
#     ws.title = f"Płyty i operacje od {start_date} do {end_date}"
#
#     # Styl nagłówka
#     header_font = Font(bold=True)
#     header_fill = PatternFill(start_color="a04000", end_color="a04000", fill_type="solid")
#     header_alignment = Alignment(horizontal="center", vertical="center")
#
#     # Styl danych
#     data_alignment = Alignment(horizontal="center", vertical="center")
#     currency_alignment = Alignment(horizontal="right", vertical="center")
#     plyta_fill = PatternFill(start_color="fcf3cf", end_color="fcf3cf", fill_type="solid")  # Blado żółty
#
#     # Styl obramowania
#     border_color = "e59866"
#     thin_border = Border(
#         left=Side(style="thin", color=border_color),
#         right=Side(style="thin", color=border_color),
#         top=Side(style="thin", color=border_color),
#         bottom=Side(style="thin", color=border_color)
#     )
#
#     # Nagłówki
#     headers = [
#         "ID Prod.", "Magazyn", "Nazwa", "Stan", "Jednostka", "Opis", "Cena",
#         "Typ Operacji", "ID Dok.", "Źródło/Cel", "Data", "Ilość", "Jednostka", "Kwota/Cena"
#     ]
#     ws.append(headers)
#
#     # Formatowanie nagłówków
#     for col_num, _ in enumerate(headers, start=1):
#         cell = ws.cell(row=1, column=col_num)
#         cell.font = header_font
#         cell.fill = header_fill
#         cell.alignment = header_alignment
#         cell.border = thin_border
#
#     # Dane
#     for plyta in plyty:
#         # Dodajemy główną linię dla Plyta
#         mag = ''
#         if plyta.magazyn == 'MAGAZYN1':
#             mag = 'SZPARAGOWA'
#         if plyta.magazyn == 'MAGAZYN2':
#             mag = 'PODOLANY'
#
#         plyta_row = [
#             plyta.prod_id,
#             mag,
#             plyta.nazwa,
#             plyta.stan,
#             plyta.jm,
#             plyta.opis,
#             f"{plyta.cena.amount} {plyta.cena.currency}" if plyta.cena else "",
#             "", "", "", "", "", "", ""
#         ]
#         ws.append(plyta_row)
#
#         # Formatowanie wiersza Plyta
#         row_num = ws.max_row
#         for col_num, value in enumerate(plyta_row, start=1):
#             cell = ws.cell(row=row_num, column=col_num)
#             # Kolumny 7 i 14: wyrównanie do prawej
#             if col_num == 7 or col_num == 14:
#                 cell.alignment = currency_alignment
#             else:
#                 cell.alignment = data_alignment
#             cell.fill = plyta_fill
#             cell.border = thin_border
#
#         # Dodajemy linie z tabeli Przychod powiązane z tą Plyta
#         plyta_przychody = przychody.filter(plyta=plyta)
#         for przychod in plyta_przychody:
#             przychod_row = [
#                 "", "", "", "", "", "", "",
#                 "Przychód",
#                 przychod.doc_id,
#                 przychod.zrodlo,
#                 przychod.data,
#                 przychod.ilosc,
#                 przychod.jm,
#                 f"{przychod.cena.amount} {przychod.cena.currency}" if przychod.cena else ""
#             ]
#             ws.append(przychod_row)
#
#             # Formatowanie wiersza Przychod
#             row_num = ws.max_row
#             for col_num, value in enumerate(przychod_row, start=1):
#                 cell = ws.cell(row=row_num, column=col_num)
#                 if col_num == 7 or col_num == 14:
#                     cell.alignment = currency_alignment
#                 else:
#                     cell.alignment = data_alignment
#                 cell.border = thin_border
#
#         # Dodajemy linie z tabeli Rozchod powiązane z tą Plyta
#         plyta_rozchody = rozchody.filter(plyta=plyta)
#         for rozchod in plyta_rozchody:
#             rozchod_row = [
#                 "", "", "", "", "", "", "",
#                 "Rozchód",
#                 rozchod.doc_id,
#                 rozchod.cel,
#                 rozchod.data,
#                 rozchod.ilosc,
#                 rozchod.jm,
#                 f"{rozchod.kwota.amount} {rozchod.kwota.currency}" if rozchod.kwota else ""
#             ]
#             ws.append(rozchod_row)
#
#             # Formatowanie wiersza Rozchod
#             row_num = ws.max_row
#             for col_num, value in enumerate(rozchod_row, start=1):
#                 cell = ws.cell(row=row_num, column=col_num)
#                 if col_num == 7 or col_num == 14:
#                     cell.alignment = currency_alignment
#                 else:
#                     cell.alignment = data_alignment
#                 cell.border = thin_border
#
#     # Automatyczne dopasowanie szerokości kolumn
#     for col in ws.columns:
#         max_length = 0
#         col_letter = col[0].column_letter  # Pobranie litery kolumny
#         for cell in col:
#             try:
#                 if cell.value:
#                     max_length = max(max_length, len(str(cell.value)))
#             except:
#                 pass
#         adjusted_width = max_length + 2
#         ws.column_dimensions[col_letter].width = adjusted_width
#
#     # Zapisanie workbooka do strumienia
#     output = BytesIO()
#     wb.save(output)
#     output.seek(0)
#
#     # Przygotowanie odpowiedzi HTTP
#     response = HttpResponse(output, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
#     response["Content-Disposition"] = f'attachment; filename="Raport_{start_date}_do_{end_date}.xlsx"'
#     return response
#





